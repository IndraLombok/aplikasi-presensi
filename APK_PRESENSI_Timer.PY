import openpyxl
from openpyxl import Workbook
import PySimpleGUI as sg
from datetime import datetime

class Presensi:
    def __init__(self):
        self.data = []
        self.filename = "presensi.xlsx"
        self.load_data_from_excel()

    def load_data_from_excel(self):
        try:
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.data.append({
                    "nama": row[0],
                    "tanggal": datetime.strptime(row[1], '%d-%m-%Y').date(),
                    "waktu": row[2],
                    "status": row[3]
                })
        except FileNotFoundError:
            self.save_data_to_excel()
        except Exception as e:
            sg.popup(f"Terjadi kesalahan saat memuat data dari Excel: {e}")

    def save_data_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nama", "Tanggal", "Waktu", "Status"])
        for presensi in self.data:
            sheet.append([presensi["nama"], presensi["tanggal"].strftime('%d-%m-%Y'), presensi["waktu"], presensi["status"]])
        workbook.save(self.filename)

    def tambah_presensi(self, nama, tanggal, waktu, status):
        try:
            tanggal = datetime.strptime(tanggal, '%d-%m-%Y').date()
            datetime.strptime(waktu, '%H:%M')
            self.data.append({"nama": nama, "tanggal": tanggal, "waktu": waktu, "status": status})
            self.save_data_to_excel()
        except ValueError:
            sg.popup("Format tanggal atau waktu tidak valid. Tanggal harus dalam format DD-MM-YYYY dan waktu dalam format HH:MM.")

    def cari_presensi(self, nama_or_tanggal):
        for presensi in self.data:
            if presensi["nama"] == nama_or_tanggal or presensi["tanggal"].strftime('%d-%m-%Y') == nama_or_tanggal:
                return presensi
        return None

    def urutkan_presensi(self):
        self.data.sort(key=lambda x: x["tanggal"])
        self.save_data_to_excel()

def is_within_time_range(start_hour, end_hour):
    current_hour = datetime.now().hour
    return start_hour <= current_hour < end_hour

def main():
    if not is_within_time_range(7, 12):  # Misalnya aplikasi hanya bisa diakses antara jam 7 pagi hingga 12 siang
        sg.popup("Aplikasi presensi hanya bisa diakses antara jam 07:00 hingga 12:00.")
        return

    presensi = Presensi()
    
    sg.theme('LightBlue2')

    layout = [
        [sg.Text('Nama'), sg.InputText(key='-NAMA-')],
        [sg.Text('Tanggal'), sg.InputText(key='-TANGGAL-', size=(25, 1)), sg.CalendarButton('Pilih Tanggal', target='-TANGGAL-', format='%d-%m-%Y')],
        [sg.Text('Waktu (hh:mm)'), sg.InputText(key='-WAKTU-')],
        [sg.Text('Status'), sg.Combo(['Hadir', 'Tidak Hadir'], key='-STATUS-')],
        [sg.Button('Tambah Presensi')],
        [sg.Text('_' * 50)],
        [sg.Text('Cari Nama/Tanggal'), sg.InputText(key='-CARI-')],
        [sg.Button('Cari Presensi')],
        [sg.Text('_' * 50)],
        [sg.Button('Lihat Presensi')],
        [sg.Button('Urutkan Presensi')],
        [sg.Button('Keluar')],
        [sg.Text('_' * 50)],
        [sg.Text('Hasil Pencarian/Lihat Presensi')],
        [sg.Multiline(size=(50, 10), key='-HASIL-')]
    ]

    window = sg.Window('Aplikasi Presensi', layout)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Keluar':
            break
        if event == 'Tambah Presensi':
            nama = values['-NAMA-']
            tanggal = values['-TANGGAL-']
            waktu = values['-WAKTU-']
            status = values['-STATUS-']
            if nama and tanggal and waktu and status:
                presensi.tambah_presensi(nama, tanggal, waktu, status)
                sg.popup('Presensi berhasil ditambahkan.')
                window['-NAMA-'].update('')
                window['-TANGGAL-'].update('')
                window['-WAKTU-'].update('')
                window['-STATUS-'].update('')
            else:
                sg.popup('Semua kolom harus diisi.')
        if event == 'Cari Presensi':
            nama_or_tanggal = values['-CARI-']
            hasil = presensi.cari_presensi(nama_or_tanggal)
            if hasil:
                hasil_text = f"Nama: {hasil['nama']}\nTanggal: {hasil['tanggal'].strftime('%d-%m-%Y')}\nWaktu: {hasil['waktu']}\nStatus: {hasil['status']}\n"
            else:
                hasil_text = 'Presensi tidak ditemukan.'
            window['-HASIL-'].update(hasil_text)
        if event == 'Lihat Presensi':
            if len(presensi.data) == 0:
                hasil_text = 'Tidak ada presensi.'
            else:
                hasil_text = ''
                for presensi_item in presensi.data:
                    hasil_text += f"Nama: {presensi_item['nama']}\nTanggal: {presensi_item['tanggal'].strftime('%d-%m-%Y')}\nWaktu: {presensi_item['waktu']}\nStatus: {presensi_item['status']}\n\n"
            window['-HASIL-'].update(hasil_text)
        if event == 'Urutkan Presensi':
            presensi.urutkan_presensi()
            sg.popup('Presensi telah diurutkan berdasarkan tanggal.')

    window.close()

if __name__ == "__main__":
    main()
